
import os
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Step 1: Access the folder
base_path = r"\\LD-295\\D_Drive\\KITE_DATA\\Report"
# base_path = r"\\Ld-324\\d\\KITE_DATA\\Report\\"
folders = [os.path.join(base_path, f) for f in os.listdir(base_path) if os.path.isdir(os.path.join(base_path, f))]
if not folders:
    raise FileNotFoundError("No folders found in the specified path.")

# Step 2: Open the latest folder
latest_folder = max(folders, key=os.path.getmtime)

# Step 3: Open MainDetailedReport.html
html_file = os.path.join(latest_folder, "MainDetailedReport.html")
if not os.path.isfile(html_file):
    raise FileNotFoundError("MainDetailedReport.html not found in the latest folder.")

# Step 4: Extract all available tables from HTML report
with open(html_file, "r", encoding="utf-8") as file:
    soup = BeautifulSoup(file, "html.parser")
tables = soup.find_all("table")
if not tables:
    raise ValueError("No tables found in the HTML file.")

# Step 5: Create Excel workbook and sheet
wb = Workbook()
ws = wb.active
ws.title = "Execution Report"

# Step 6: Clear A1 (no merge)
ws["A1"].value = None

# Define fill colors
fill_colors = {
    "passed": PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid"),
    "failed": PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"),
    "not executed": PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"),
    "verify manually": PatternFill(start_color="FFA52A2A", end_color="FFA52A2A", fill_type="solid")
}

# Define styles
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
bold_font = Font(name="Calibri", size=11, bold=True)
default_font = Font(name="Calibri", size=11)
center_alignment = Alignment(horizontal="center", vertical="center")

# Helper function to determine fill color
def get_fill_color(row_text):
    row_text_lower = row_text.lower()
    if "passed" in row_text_lower:
        return fill_colors["passed"]
    elif "failed" in row_text_lower:
        return fill_colors["failed"]
    elif "not executed" in row_text_lower:
        return fill_colors["not executed"]
    elif "verify manually" in row_text_lower:
        return fill_colors["verify manually"]
    return None

# Helper function to write a table to the worksheet
def write_table_to_sheet(table, start_row):
    rows = table.find_all("tr")
    for i, row in enumerate(rows):
        cols = row.find_all(["th", "td"])
        row_text = row.get_text(strip=True)
        fill = get_fill_color(row_text)
        for j, col in enumerate(cols):
            cell = ws.cell(row=start_row + i, column=j + 1, value=col.get_text(strip=True))
            cell.font = bold_font if row.find("th") else default_font
            cell.alignment = center_alignment
            cell.border = thin_border
            if fill:
                cell.fill = fill
    return start_row + len(rows)

# Write all tables sequentially with a blank row in between
next_row = 2
for table in tables:
    next_row = write_table_to_sheet(table, start_row=next_row)
    next_row += 1

# Apply header coloring based on column names
header_row = ws[11]
for cell in header_row:
    if cell.value:
        val = cell.value.lower()
        if "passed" in val:
            cell.fill = fill_colors["passed"]
        elif "failed" in val:
            cell.fill = fill_colors["failed"]
        elif "not executed" in val:
            cell.fill = fill_colors["not executed"]
        elif "verify manually" in val:
            cell.fill = fill_colors["verify manually"]



 #Define fill colors
fill_colors = {
    "D": PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid"),  # Green
    "E": PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"),  # Red
    "F": PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"),  # Yellow
    "G": PatternFill(start_color="FFA52A2A", end_color="FFA52A2A", fill_type="solid")   # Brown
}

# Remove all fill colors from row 12
for cell in ws[12]:
    cell.fill = PatternFill(fill_type=None)

# Dynamically apply fill colors to D-G from row 12 to last row
max_row = ws.max_row
for row in range(12, max_row + 1):
    for col_letter in ['D', 'E', 'F', 'G']:
        cell = ws[f"{col_letter}{row}"]
        if cell.value not in [None, ""]:
            cell.fill = fill_colors[col_letter]



# Auto-adjust column widths
for col in ws.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

# Merge A2 and B2
ws.merge_cells("A2:B2")

# Column J is the 10th column
ws.delete_cols(10)  

# Save Excel file
excel_path = os.path.join(latest_folder, "MainDetailedReport_formatted.xlsx")
wb.save(excel_path)

print(f"✅ Excel file saved at: {excel_path}")

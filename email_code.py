
import os
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from win32com.client import gencache

# ------------------- CONFIGURABLE PARAMETERS -------------------
build_number = "KITE_5.0.1.25162.13"
machine_name = "LD 256"
recipient = "Somesh"
workspace_url = r"\\192.x.x.x\\Kite_Drive\\KITE_DATA\\Execution\\ExecutionReport_01_08_12_17_48_41"
overnight_report_path = r"\\192.x.x.x\\Kite_Drive\\KITE_DATA\\Execution\\ExecutionReport_01_08_12_17, 48, 41"
percentage_report_path = r"\\192.x.x.x\\Drive_KITE_DATA\\ResourceExecutionReport_O1_O8_I2_I7, 48, 41"
# ---------------------------------------------------------------

base_path = r"\\LD-295\\D_Drive\\KITE_DATA\\Report"
folders = [os.path.join(base_path, f) for f in os.listdir(base_path) if os.path.isdir(os.path.join(base_path, f))]
latest_folder = max(folders, key=os.path.getmtime)
html_file = os.path.join(latest_folder, "MainDetailedReport.html")

with open(html_file, "r", encoding="utf-8") as file:
    soup = BeautifulSoup(file, "html.parser")
tables = soup.find_all("table")

wb = Workbook()
ws = wb.active
ws.title = "Execution Report"
ws["A1"].value = None

fill_colors = {
    "passed": PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid"),
    "failed": PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"),
    "not executed": PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"),
    "verify manually": PatternFill(start_color="FF7B3F00", end_color="FF7B3F00", fill_type="solid")
}
dg_fill_colors = {
    "D": PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid"),
    "E": PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"),
    "F": PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"),
    "G": PatternFill(start_color="FFA52A2A", end_color="FFA52A2A", fill_type="solid")
}
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
bold_font = Font(name="Calibri", size=11, bold=True)
default_font = Font(name="Calibri", size=11)
center_alignment = Alignment(horizontal="center", vertical="center")

def get_fill_color(row_text):
    row_text_lower = row_text.lower()
    for key in fill_colors:
        if key in row_text_lower:
            return fill_colors[key]
    return None

def write_table_to_sheet(table, start_row):
    rows = table.find_all("tr")
    for i, row in enumerate(rows):
        cols = row.find_all(["th", "td"])
        row_text = row.get_text(strip=True)
        fill = get_fill_color(row_text)
        for j, col in enumerate(cols):
            cell = ws.cell(row=start_row + i, column=j + 1, value=col.get_text(strip=True))
            cell.font = bold_font if row.find("th") else default_font
            cell.alignment = center_alignment
            cell.border = thin_border
            if fill:
                cell.fill = fill
    return start_row + len(rows)

next_row = 2
table_start_rows = []
for table in tables:
    table_start_rows.append(next_row)
    next_row = write_table_to_sheet(table, start_row=next_row)
    next_row += 1

for cell in ws[12]:
    cell.fill = PatternFill(fill_type=None)

max_row = ws.max_row
for row in range(12, max_row + 1):
    for col_letter in ['D', 'E', 'F', 'G']:
        cell = ws[f"{col_letter}{row}"]
        if cell.value not in [None, ""]:
            cell.fill = dg_fill_colors[col_letter]

for col in ws.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

excel_path = os.path.join(latest_folder, "MainDetailedReport_formatted.xlsx")
wb.save(excel_path)

def excel_tables_to_html(excel_path, sheet_name="Execution Report", table_start_rows=None):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    html_tables = []
    for start in table_start_rows:
        rows = []
        for row in ws.iter_rows(min_row=start, max_row=start+10, values_only=True):
            row_html = "".join(f"<td>{str(cell) if cell else ''}</td>" for cell in row)
            rows.append(f"<tr>{row_html}</tr>")
        html_table = "<table border='1' cellpadding='4' cellspacing='0' style='border-collapse:collapse;font-family:Calibri;font-size:11pt;'>" + "".join(rows) + "</table>"
        html_tables.append(html_table)
    return html_tables

def compose_email_body(build_number, machine_name, workspace_url, overnight_report_path, percentage_report_path, html_tables):
    body = f"""
    <p>Hi Team,</p>
    <p>Please find below the Overnight Execution Report on Machine <b>{machine_name}</b> for Build <b>{build_number}</b> on QN Resource Workspace Command line with UI with Monitoring Microsoft Camera.</p>
    <p>
    Workspace URL: {workspace_url}<br>
    Path Overnight Execution Report: {overnight_report_path}<br>
    Path Percentage Execution Report: {percentage_report_path}
    </p>
    <p><b>Execution Summary:</b></p>
    {html_tables[0]}
    <p><b>Detailed Report:</b></p>
    {html_tables[1]}
    <p>Thanks & Regards,<br>Somesh Rajappa</p>
    """
    return body

def send_email_with_tables(subject, body_html, to):
    outlook = gencache.EnsureDispatch('Outlook.Application')
    mail = outlook.CreateItem(0)
    mail.To = to
    mail.Subject = subject
    mail.HTMLBody = body_html
    mail.Send()

html_tables = excel_tables_to_html(excel_path, table_start_rows=table_start_rows)
subject = f"Overnight Report Execution Status Report for Build {build_number}"
body_html = compose_email_body(build_number, machine_name, workspace_url, overnight_report_path, percentage_report_path, html_tables)
send_email_with_tables(subject, body_html, recipient)

print("✅ Outlook email sent with summary and detailed tables in the body.")
